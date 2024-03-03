import openpyxl

# open xl(workbook)-->  sheets--> row and col
path = r"C:\Users\user\PycharmProjects\pythonProject_WE_October\test_framework\test_data_with_xl\simple_interest.xlsx"
wb = openpyxl.load_workbook(path)
print(wb)
print(wb.sheetnames)
print(wb.active)

sheet=wb["Sheet1"]

print(sheet.max_row)
print(sheet.max_column)
####read
# for i in range(2,sheet.max_row+1):#(0 to n-1)
#     for j in range(1,sheet.max_column+1):
#         print(sheet.cell(i, j).value, end="\t")
#     print("\n")

####write

sheet.cell(1, 10).value="heading"
sheet.cell(2, 10).value="abc"
sheet.cell(3, 10).value="abc"
sheet.cell(4, 10).value="abc"
sheet.cell(5, 10).value="abc"
sheet.cell(6, 10).value="abc"
sheet.cell(7, 10).value="abc"

wb.save(path)

#
