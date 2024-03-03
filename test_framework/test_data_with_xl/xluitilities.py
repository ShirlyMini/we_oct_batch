import openpyxl
from openpyxl.styles import PatternFill

def no_rows(xl_path, sheet):
    wb = openpyxl.load_workbook(xl_path)
    sheet = wb[sheet]
    return sheet.max_row

def no_col(xl_path, sheet):
    wb = openpyxl.load_workbook(xl_path)
    sheet = wb[sheet]
    return sheet.max_column

def read_cell(xl_path, sheet, r, c):
    wb = openpyxl.load_workbook(xl_path)
    sheet = wb[sheet]
    return sheet.cell(r,c).value

def write_cell(xl_path, sheet, r, c, data):
    wb = openpyxl.load_workbook(xl_path)
    sheet = wb[sheet]
    sheet.cell(r, c).value = data
    wb.save(xl_path)

def green_pattern(xl_path, sheet, r, c):
    wb = openpyxl.load_workbook(xl_path)
    sheet = wb[sheet]
    green = PatternFill(start_color="46eb34", end_color="46eb34", fill_type="solid")
    sheet.cell(r, c).fill = green
    wb.save(xl_path)


def red_pattern(xl_path, sheet, r, c):
    wb = openpyxl.load_workbook(xl_path)
    sheet = wb[sheet]
    red = PatternFill(start_color="eb3434", end_color="eb3434", fill_type="solid")
    sheet.cell(r, c).fill = red
    wb.save(xl_path)